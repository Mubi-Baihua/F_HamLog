def main(file):
    from PySide6.QtWidgets import QFileDialog, QMessageBox, QApplication
    from dialog_defaults import desktop_dir
    import os
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # 获取父窗口：优先用当前活动窗口（用户正在操作的那一个），
    # 避免导出对话框把主窗口强行提到最前层（例如从“搜索结果”窗口导出时）。
    parent_window = None
    app = QApplication.instance()
    if app:
        active = app.activeWindow()
        if active is not None and active.isVisible():
            parent_window = active
        else:
            for window in app.topLevelWidgets():
                if window.isVisible() and window.windowTitle().startswith('F HamLog'):
                    parent_window = window
                    break

    # 已知字段的中文列名（按常用顺序；不含 record，避免超大 base64 写入表格）
    translation_dict = {
        'date': '日期',
        'time': '时间',
        'm_call': '己方呼号',
        'o_call': '对方呼号',
        'freq': '频率',
        'freq_rx': '下行频率',
        'mode': '调制模式',
        'prop_mode': '传播模式',
        'sat_name': '卫星名称',
        'm_rst': '己方接收信号',
        'o_rst': '对方接收信号',
        'm_qth': '己方QTH',
        'o_qth': '对方QTH',
        "m_dig": '己方设备',
        'o_dig': '对方设备',
        'm_ant': '己方天线',
        'o_ant': '对方天线',
        'm_pow': '己方功率',
        'o_pow': '对方功率',
        'notes': '备注'
    }

    # 动态收集数据中实际出现的全部字段（排除 record），保证不遗漏任何日志条目字段
    seen_keys = []
    for entry in (file or []):
        if isinstance(entry, dict):
            for k in entry.keys():
                if k != 'record' and k not in seen_keys:
                    seen_keys.append(k)
    # 列顺序：已知字段优先（按其定义顺序），其余未知字段按出现顺序追加在后
    ordered_keys = [k for k in translation_dict if k in seen_keys]
    for k in seen_keys:
        if k not in ordered_keys:
            ordered_keys.append(k)
    headers = [translation_dict.get(k, k) for k in ordered_keys]
    keys = ordered_keys

    # 弹出保存对话框
    save_path, _ = QFileDialog.getSaveFileName(
        parent_window,
        "导出为Excel文件",
        desktop_dir(),
        "Excel文件 (*.xlsx);;所有文件 (*)"
    )
    if not save_path:
        return False
    if not save_path.lower().endswith('.xlsx'):
        save_path += '.xlsx'

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = '通联日志'

        # 写入表头
        ws.append(headers)

        # 逐行写入数据（entry 可能不是字典，做容错）
        for entry in (file or []):
            if not isinstance(entry, dict):
                entry = {}
            ws.append([entry.get(k, '') for k in keys])

        # 自动调整列宽（仅遍历有数据的列/行）
        for col_idx in range(1, len(keys) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=col_idx).value
                if cell_val is not None:
                    length = len(str(cell_val))
                    if length > max_len:
                        max_len = length
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

        wb.save(save_path)
        QMessageBox.information(
            parent_window, "导出成功",
            f"已导出 {len(file or [])} 条记录到：\n{save_path}")
        return True

    except Exception as e:
        if parent_window:
            QMessageBox.critical(parent_window, "导出失败", f"导出过程中发生错误：\n{str(e)}")
        else:
            print(f"导出过程中发生错误：{str(e)}")
        return False
