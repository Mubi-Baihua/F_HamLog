from openpyxl import load_workbook
from PySide6.QtWidgets import QFileDialog
from dialog_defaults import desktop_dir
from PySide6.QtWidgets import *
from datetime import datetime, timedelta, timezone
import re


def _utc_cell_to_local(date_time_cell):
    # 支持 openpyxl 返回的 datetime 或字符串形式 'YYYY-MM-DDTHH:MM' 或 'YYYYMMDDTHHMM'
    if date_time_cell is None:
        return None, None
    if isinstance(date_time_cell, datetime):
        utc_dt = date_time_cell if date_time_cell.tzinfo else date_time_cell.replace(tzinfo=timezone.utc)
        local_tz = datetime.now().astimezone().tzinfo
        local_dt = utc_dt.astimezone(local_tz)
        return local_dt.strftime('%Y-%m-%d'), local_dt.strftime('%H:%M')

    s = str(date_time_cell).strip()
    if 'T' in s:
        parts = s.split('T')
        if len(parts) >= 2:
            date_part = parts[0]
            time_part = parts[1]
        else:
            return None, None
    else:
        # 尝试从连体字符串中切分
        m = re.match(r"(\d{8}).*(\d{4})", s)
        if m:
            date_part = m.group(1)
            time_part = m.group(2)
        else:
            return None, None

    # 规范化
    date_part = date_part.replace('-', '')
    time_part = re.sub(r'[^0-9]', '', time_part)
    if len(date_part) < 8 or len(time_part) < 4:
        return None, None

    try:
        utc_dt = datetime(int(date_part[0:4]), int(date_part[4:6]), int(date_part[6:8]), int(time_part[0:2]), int(time_part[2:4]), tzinfo=timezone.utc)
    except Exception:
        return None, None

    local_tz = datetime.now().astimezone().tzinfo
    local_dt = utc_dt.astimezone(local_tz)
    return local_dt.strftime('%Y-%m-%d'), local_dt.strftime('%H:%M')


def main(file):
    print("导入HAM_tolls")

    file_path, _ = QFileDialog.getOpenFileName(None, "选择Excel文件", desktop_dir(), "Excel文件 (*.xlsx *.xls)")

    if not file_path:
        return file

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    total_rows = worksheet.max_row

    new_records = []
    for row in range(2, total_rows + 1):  # 假设第一行是标题，从第二行开始读取
        rec = {
            'date': '',
            'time': '',
            'm_call': '',
            'o_call': '',
            'freq': '',
            'mode': '',
            'm_rst': '',
            'o_rst': '',
            'm_qth': '',
            'o_qth': '',
            "m_dig": '',
            'o_dig': '',
            'm_ant': '',
            'o_ant': '',
            'm_pow': '',
            'o_pow': '',
            'notes': ''
        }

        cell_val = worksheet[f'B{row}'].value
        local_date, local_time = _utc_cell_to_local(cell_val)
        if local_date and local_time:
            rec['date'] = local_date
            rec['time'] = local_time
        else:
            # 回退：尝试原先的字符串拆分并按东八区转换（兼容旧数据）
            try:
                s = str(cell_val)
                if 'T' in s:
                    dpart, tpart = s.split('T')[0], s.split('T')[1]
                    rec['date'] = dpart
                    utc_time_str = tpart[0:5]
                    utc_time = datetime.strptime(utc_time_str, "%H:%M")
                    beijing_time = utc_time + timedelta(hours=8)
                    rec['time'] = beijing_time.strftime("%H:%M")
            except Exception:
                rec['date'] = ''
                rec['time'] = ''

        rec['m_call'] = worksheet[f'C{row}'].value or ''
        rec['o_call'] = worksheet[f'D{row}'].value or ''
        rec['freq'] = worksheet[f'I{row}'].value or ''
        rec['mode'] = worksheet[f'J{row}'].value or ''
        rec['m_rst'] = worksheet[f'G{row}'].value or ''
        rec['o_rst'] = worksheet[f'H{row}'].value or ''
        rec['m_qth'] = worksheet[f'E{row}'].value or ''
        rec['o_qth'] = worksheet[f'F{row}'].value or ''
        rec['m_dig'] = worksheet[f'K{row}'].value or ''
        rec['o_dig'] = worksheet[f'L{row}'].value or ''
        rec['notes'] = worksheet[f'M{row}'].value or ''

        new_records.append(rec)

    # 重复检测：按 o_call + date + time (均小写/strip)
    out = list(file)
    existing_keys = set()
    for e in out:
        key = ((e.get('o_call') or '').strip().lower(), e.get('date', ''), e.get('time', ''))
        if key[0] and key[1] and key[2]:
            existing_keys.add(key)

    mapped_dup = []
    mapped_nondup = []
    dup_keys = set()
    for r in new_records:
        key = ((r.get('o_call') or '').strip().lower(), r.get('date', ''), r.get('time', ''))
        if key[0] and key[1] and key[2] and key in existing_keys:
            mapped_dup.append(r)
            dup_keys.add(key)
        else:
            mapped_nondup.append(r)

    if mapped_dup:
        msg = QMessageBox()
        msg.setWindowTitle("检测到重复记录")
        msg.setText(f"检测到 {len(mapped_dup)} 条导入记录与现有记录重复。请选择处理方式：")
        overwrite_btn = msg.addButton("全部覆盖", QMessageBox.AcceptRole)
        skip_btn = msg.addButton("全部跳过", QMessageBox.RejectRole)
        msg.exec()
        if msg.clickedButton() == overwrite_btn:
            out = [e for e in out if ((e.get('o_call') or '').strip().lower(), e.get('date', ''), e.get('time', '')) not in dup_keys]
            out.extend(new_records)
        else:
            out.extend(mapped_nondup)
    else:
        out.extend(new_records)

    return out
if __name__ == '__main__':
    app = QApplication()
    print(main([]))