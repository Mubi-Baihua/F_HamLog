from openpyxl import load_workbook
from PySide6.QtWidgets import QFileDialog
from PySide6.QtWidgets import *
from datetime import datetime, timedelta
def main(file):
    print("导入HAM_tolls")

    
    file_path, _ = QFileDialog.getOpenFileName(
        None, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
    )
    
    if file_path:
        # 读取Excel文件
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        total_rows = worksheet.max_row
        for row in range(2, total_rows + 1):  # 假设第一行是标题，从第二行开始读取
            file.append({
            'date': '',
            'time': '',
            'm_call': '',
            'o_call': '',
            'freq': '',
            'mode': '',
            'm_rst': '',
            'o_rst': '',
            'm_qth': '',
            'o_qth': '',
            "m_dig": '',
            'o_dig': '',
            'm_ant': '',
            'o_ant': '',
            'm_pow': '',
            'o_pow': '',
            'notes': ''
            })
            file[-1]['date'] = worksheet[f'B{row}'].value.split('T')[0]

            utc_time_str = worksheet[f'B{row}'].value.split('T')[1][0:5]
            # 将UTC时间转换为北京时间
            utc_time = datetime.strptime(utc_time_str, "%H:%M")
            beijing_time = utc_time + timedelta(hours=8)
            file[-1]['time'] = beijing_time.strftime("%H:%M")

            file[-1]['m_call'] = worksheet[f'C{row}'].value
            file[-1]['o_call'] = worksheet[f'D{row}'].value
            file[-1]['freq'] = worksheet[f'I{row}'].value
            file[-1]['mode'] = worksheet[f'J{row}'].value
            file[-1]['m_rst'] = worksheet[f'G{row}'].value
            file[-1]['o_rst'] = worksheet[f'H{row}'].value
            file[-1]['m_qth'] = worksheet[f'E{row}'].value
            file[-1]['o_qth'] = worksheet[f'F{row}'].value
            file[-1]['m_dig'] = worksheet[f'K{row}'].value
            file[-1]['o_dig'] = worksheet[f'L{row}'].value
            file[-1]['notes'] = worksheet[f'M{row}'].value

    return file
if __name__ == '__main__':
    app = QApplication()
    print(main([]))